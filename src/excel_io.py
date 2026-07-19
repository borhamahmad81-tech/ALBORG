"""
excel_io.py

Reads the input patient list (Patient ID + Name) and writes the master
output workbook with one row per lab test result (long format), plus an
Errors sheet listing any patient/report that failed to fetch or parse.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@dataclass
class PatientEntry:
    patient_id: str
    name: str


def read_patient_list(path: str, id_column: str = "Patient No",
                       name_column: str = "Name") -> list[PatientEntry]:
    """Read the input Excel sheet. Column names are matched case-insensitively
    and with flexible spacing, so 'Patient No', 'Patient ID', 'PatientNo' all
    work. Only the ID is required for searching; Name is used for labeling."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            key = str(cell.value).strip().lower().replace(" ", "").replace("_", "")
            headers[key] = col_idx

    def find_col(*candidates):
        for c in candidates:
            key = c.strip().lower().replace(" ", "").replace("_", "")
            if key in headers:
                return headers[key]
        return None

    id_col = find_col(id_column, "patientid", "patientno", "id", "mrn")
    name_col = find_col(name_column, "patientname", "fullname")

    if id_col is None:
        raise ValueError(
            f"Could not find a patient ID column in '{path}'. "
            f"Found columns: {list(headers.keys())}"
        )

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        pid_cell = row[id_col - 1]
        if pid_cell.value is None or str(pid_cell.value).strip() == "":
            continue
        pid = str(pid_cell.value).strip()
        # Excel sometimes stores numeric IDs as floats (e.g. 1036240040.0)
        if pid.endswith(".0"):
            pid = pid[:-2]
        name = ""
        if name_col is not None:
            name_val = row[name_col - 1].value
            name = str(name_val).strip() if name_val is not None else ""
        entries.append(PatientEntry(patient_id=pid, name=name))

    return entries


RESULT_HEADERS = [
    "Patient ID", "Patient Name", "Test", "Result", "Reported On",
]

ERROR_HEADERS = ["Patient ID", "Patient Name", "Stage", "Details"]


RESULT_KEYS = ["patient_id", "patient_name", "test", "result", "reported_on"]
ERROR_KEYS = ["patient_id", "patient_name", "stage", "details"]
UNPARSED_KEYS = ["patient_id", "patient_name", "line"]


def read_existing_progress(path: str) -> tuple[list[dict], list[dict], list[dict]]:
    """Read back a previous run's output (if it exists) so a rerun can resume
    instead of starting from scratch. Returns (results, errors, unparsed) as
    lists of dicts, matching the shape main.py builds them in. Old error rows
    are intentionally dropped - patients not yet successfully completed will
    simply be retried and get fresh errors if they fail again."""
    if not Path(path).exists():
        return [], [], []

    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return [], [], []

    def read_sheet(name, keys):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            rows.append({k: ("" if v is None else v) for k, v in zip(keys, row)})
        return rows

    results = read_sheet("Lab Results", RESULT_KEYS)
    unparsed = read_sheet("Unparsed Lines", UNPARSED_KEYS)
    return results, [], unparsed


def get_completed_patient_ids(results: list[dict]) -> set[str]:
    return {str(r["patient_id"]) for r in results if r.get("patient_id")}


def write_master_workbook(output_path: str, all_results: list[dict],
                           errors: list[dict], unparsed: list[dict]) -> None:
    """
    all_results: list of dicts matching RESULT_HEADERS keys (snake_case)
    errors: list of dicts with keys: patient_id, patient_name, stage, details
    unparsed: list of dicts with keys: patient_id, patient_name, line
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Lab Results"
    _write_sheet(ws, RESULT_HEADERS, [
        [
            r.get("patient_id", ""), r.get("patient_name", ""),
            r.get("test", ""), _numeric_result(r.get("result", "")),
            r.get("reported_on", ""),
        ]
        for r in all_results
    ])

    ws_err = wb.create_sheet("Errors")
    _write_sheet(ws_err, ERROR_HEADERS, [
        [e.get("patient_id", ""), e.get("patient_name", ""),
         e.get("stage", ""), e.get("details", "")]
        for e in errors
    ])

    if unparsed:
        ws_unparsed = wb.create_sheet("Unparsed Lines")
        _write_sheet(ws_unparsed, ["Patient ID", "Patient Name", "Line"], [
            [u.get("patient_id", ""), u.get("patient_name", ""), u.get("line", "")]
            for u in unparsed
        ])

    wb.save(output_path)


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] + [len(str(row[col_idx - 1])) for row in rows]
        ) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"


def _numeric_result(value) -> object:
    """Return the result as a real number so Excel can sort/average/chart it.

    Values that genuinely aren't plain numbers (e.g. '<0.01', '>1000',
    'Positive') are left as text - forcing those into a number would
    silently change their meaning."""
    if value is None:
        return ""
    text = str(value).strip().replace(",", "")
    if text == "":
        return ""
    try:
        num = float(text)
    except ValueError:
        return text  # e.g. '<0.01', 'Positive' - keep exactly as reported
    # Use int when there's no fractional part, so 106.0 shows as 106.
    return int(num) if num.is_integer() and "." not in text else num
